from data.models.pokemon_set_model import PokemonSet
from data.models.pokemon_column_model import PokeColumn
import openpyxl
from openpyxl import worksheet
from openpyxl import workbook

DEBUG_MODE = False


def get_poke_columns_config():
    return [PokeColumn('Card #', 2),
            PokeColumn('# Owned', 3),
            PokeColumn('Name', 1),
            PokeColumn('Rarity', 4),
            PokeColumn('Type', 5)]


def _get_excel_workbook_from_file(file_path: str) -> workbook:
    return openpyxl.load_workbook(file_path)


class PokemonSetSheet:
    def __init__(self, pokemon_set: PokemonSet, excel_workbook: workbook, excel_sheet: worksheet, file_path: str):
        self.pokemon_set = pokemon_set
        self.excel_workbook: openpyxl.workbook = excel_workbook
        self.excel_sheet: worksheet = excel_sheet
        self.file_path = file_path
        self.column_config = get_poke_columns_config()
        self.__column_offset__ = self.column_config.__len__() + self.excel_sheet.max_column
        # Perform setup util
        self.configure_columns()

    def _move_column_from_index_to_other_index(self, index, other_index):
        values_to_move = []
        for i in range(1, self.excel_sheet.max_row + 1):
            values_to_move.append(self.excel_sheet.cell(row=i, column=index).value)
            self.excel_sheet.cell(row=i, column=index).value = ''

        for j in range(1, self.excel_sheet.max_row + 1):
            self.excel_sheet.cell(row=j, column=other_index).value = values_to_move[j-1]

    def _set_cell_value_at(self, value, row, column):
        self.excel_sheet.cell(row=row, column=column).value = value

    def _is_cell_empty_at(self, row, column):
        cell_value: str = self.excel_sheet.cell(row=row, column=column).value
        return cell_value is None or str(cell_value).strip() == ''

    def is_poke_column_in_columns(self, poke_column: PokeColumn):
        for i in range(1, self.excel_sheet.max_column + 1):
            my_column = PokeColumn(self.excel_sheet.cell(row=1, column=i).value, i)
            if poke_column.equals(my_column):
                return True
        return False

    def get_column_index_with_name(self, column_name):
        for i in range(1, self.excel_sheet.max_column + 1):
            name_at_i = self.excel_sheet.cell(row=1, column=i).value
            if name_at_i == column_name:
                return i
        return -1

    def is_column_empty(self, column_index):
        for i in range(2, self.excel_sheet.max_row + 1):
            if self.excel_sheet.cell(row=i, column=column_index).value is not None:
                return False
        return True

    def configure_columns(self):
        if not DEBUG_MODE:
            self.move_existing_columns_out_of_way()
            self.move_existing_columns_to_proper_index()
            self.insert_missing_columns()

    def insert_missing_columns(self):
        for i in range(0, self.column_config.__len__()):
            config_col: PokeColumn = self.column_config[i]
            if not self.is_poke_column_in_columns(config_col):
                self.excel_sheet.cell(row=1, column=config_col.index).value = config_col.name

    def move_existing_columns_to_proper_index(self):
        for i in range(0, self.column_config.__len__()):
            config_col: PokeColumn = self.column_config[i]
            existing_col_index = self.get_column_index_with_name(config_col.name)
            if existing_col_index != -1:
                self._move_column_from_index_to_other_index(existing_col_index, config_col.index)

    def move_existing_columns_out_of_way(self):
        last_column = self.excel_sheet.max_column + 1
        for i in range(1, last_column):
            if not self.is_column_empty(i):
                self._move_column_from_index_to_other_index(i, i + self.__column_offset__)

    def row_contains_empty_cells_under_columns_in_config(self, row_index: int, columns_to_exclude=None) -> bool:
        if columns_to_exclude is None:
            columns_to_exclude = []
        for i in range(0, self.column_config.__len__()):
            poke_column = self.column_config[i]
            if poke_column.name not in columns_to_exclude and \
                    self._is_cell_empty_at(row=row_index, column=poke_column.index):
                return True
        return False

    def get_card_numbers_with_missing_data(self) -> []:
        card_numbers_with_missing_data = []
        card_number_column = self.get_column_index_with_name('Card #')
        columns_to_exclude = ['Card #', '# Owned']
        for i in range(2, self.excel_sheet.max_row + 1):
            if self.row_contains_empty_cells_under_columns_in_config(i, columns_to_exclude):
                card_numbers_with_missing_data.append(self.excel_sheet.cell(row=i, column=card_number_column).value)
        return card_numbers_with_missing_data

    def get_row_index_from_cell_value_and_column_index(self, cell_value, column_index: int) -> int:
        for i in range(2, self.excel_sheet.max_row + 1):
            existing_value = self.excel_sheet.cell(row=i, column=column_index).value
            if existing_value == cell_value:
                return i
        return -1

    def get_row_index_for_card_number(self, card_number) -> int:
        card_number_column_index = self.get_column_index_with_name('Card #')
        return self.get_row_index_from_cell_value_and_column_index(card_number, card_number_column_index)

    def get_card_number_for_row_index(self, row_index):
        card_number_column_index = self.get_column_index_with_name('Card #')
        return self.excel_sheet.cell(row=row_index, column=card_number_column_index).value

    def get_card_numbers_in_sheet(self):
        numbers = []
        for i in range(2, self.excel_sheet.max_row + 1):
            numbers.append(self.get_card_number_for_row_index(i))
        return numbers

    def update_cell_with_card_number_and_column_name(self, value, card_number: int, column_name: str):
        column_index = self.get_column_index_with_name(column_name)
        row_index_for_card_number = self.get_row_index_for_card_number(card_number)
        if row_index_for_card_number != -1:
            self._set_cell_value_at(value=value, row=row_index_for_card_number, column=column_index)

    def update_card_name_with_card_number(self, name: str, card_number: int):
        self.update_cell_with_card_number_and_column_name(value=name, card_number=card_number, column_name='Name')

    def update_rarity_with_card_number(self, rarity: str, card_number: int):
        self.update_cell_with_card_number_and_column_name(value=rarity, card_number=card_number, column_name='Rarity')

    def insert_card_number(self, card_number):
        last_row = self.excel_sheet.max_row
        card_number_column_index = self.get_column_index_with_name('Card #')
        self._set_cell_value_at(card_number, last_row + 1, card_number_column_index)

    def save(self):
        self.excel_workbook.save(self.file_path)

    def create(pokemon_set: PokemonSet, file_path: str):
        excel_workbook = _get_excel_workbook_from_file(file_path)
        excel_sheet = excel_workbook.get_sheet_by_name(pokemon_set.abbreviation)
        return PokemonSetSheet(pokemon_set, excel_workbook, excel_sheet, file_path)
