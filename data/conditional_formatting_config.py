# Author: Jacques
# Date: 12/28/2021
# Time: 12:04 PM
from openpyxl.worksheet import worksheet
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from data.poke_column_config import get_poke_columns_config
from data.utility import column_number_to_letter

COLUMNS = get_poke_columns_config()


def make_formula_text(column_letter: str, energy_type: str):
    return '=($%s1="%s")' % (column_letter, energy_type.capitalize())


def get_energy_type_column_letter():
    for column in COLUMNS:
        if column.name == 'Type':
            return column_number_to_letter(column.index)


def get_number_owned_column_letter():
    for column in COLUMNS:
        if column.name == '# Owned':
            return column_number_to_letter(column.index)


def get_last_column_index():
    last_column_index = 0
    for column in COLUMNS:
        last_column_index = max(last_column_index, column.index)
    return last_column_index


def get_applies_to_all_rows():
    last_column_index = get_last_column_index()
    last_column_letter = column_number_to_letter(last_column_index)
    return '$A2:$%s500' % last_column_letter


ENERGY_TYPE_COLUMN_LETTER = get_energy_type_column_letter()
APPLIES_TO_ALL_ROWS = get_applies_to_all_rows()

FIRE_FILL = PatternFill(bgColor='FBD1D1', fill_type='solid')
FIRE_BORDER_COLOR = Color('F75F5F', type='hex')
FIRE_BORDER_SIDE = Side(style='medium', color=FIRE_BORDER_COLOR, border_style='thin')
FIRE_BORDER = Border(left=FIRE_BORDER_SIDE,
                     right=FIRE_BORDER_SIDE,
                     top=FIRE_BORDER_SIDE,
                     bottom=FIRE_BORDER_SIDE)
FIRE_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Fire')

FIRE_FORMULA = FormulaRule(formula=[FIRE_FORMULA_TEXT], border=FIRE_BORDER, fill=FIRE_FILL)

METAL_FILL = PatternFill(bgColor='A6A6A6', fill_type='solid')
METAL_FONT = Font(color='FFFFFF')
METAL_BORDER_COLOR = Color('F2F2F2', type='hex')
METAL_BORDER_SIDE = Side(style='medium', color=METAL_BORDER_COLOR, border_style='thin')
METAL_BORDER = Border(left=METAL_BORDER_SIDE,
                     right=METAL_BORDER_SIDE,
                     top=METAL_BORDER_SIDE,
                     bottom=METAL_BORDER_SIDE)
METAL_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Metal')

METAL_FORMULA = FormulaRule(formula=[METAL_FORMULA_TEXT],
                            font=METAL_FONT,
                            border=METAL_BORDER,
                            fill=METAL_FILL)

COLORLESS_FILL = PatternFill(bgColor='F2F2F2', fill_type='solid')
COLORLESS_BORDER_COLOR = Color('BFBFBF', type='hex')
COLORLESS_BORDER_SIDE = Side(style='medium', color=COLORLESS_BORDER_COLOR, border_style='thin')
COLORLESS_BORDER = Border(left=COLORLESS_BORDER_SIDE,
                     right=COLORLESS_BORDER_SIDE,
                     top=COLORLESS_BORDER_SIDE,
                     bottom=COLORLESS_BORDER_SIDE)
COLORLESS_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Colorless')

COLORLESS_FORMULA = FormulaRule(formula=[COLORLESS_FORMULA_TEXT], border=COLORLESS_BORDER, fill=COLORLESS_FILL)

TRAINER_FILL = PatternFill(bgColor='A6A6A6', fill_type='solid')
TRAINER_FONT = Font(color='1F3214')
TRAINER_BORDER_COLOR = Color('F2F2F2', type='hex')
TRAINER_BORDER_SIDE = Side(style='medium', color=TRAINER_BORDER_COLOR, border_style='thin')
TRAINER_BORDER = Border(left=TRAINER_BORDER_SIDE,
                     right=TRAINER_BORDER_SIDE,
                     top=TRAINER_BORDER_SIDE,
                     bottom=TRAINER_BORDER_SIDE)
TRAINER_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'N/A')

TRAINER_FORMULA = FormulaRule(formula=[TRAINER_FORMULA_TEXT],
                              font=TRAINER_FONT,
                              border=TRAINER_BORDER,
                              fill=TRAINER_FILL)

DRAGON_FILL = PatternFill(bgColor='E1C665', fill_type='solid')
DRAGON_FONT = Font(color='191505')
DRAGON_BORDER_COLOR = Color('F1E4B5', type='hex')
DRAGON_BORDER_SIDE = Side(style='medium', color=DRAGON_BORDER_COLOR, border_style='thin')
DRAGON_BORDER = Border(left=DRAGON_BORDER_SIDE,
                     right=DRAGON_BORDER_SIDE,
                     top=DRAGON_BORDER_SIDE,
                     bottom=DRAGON_BORDER_SIDE)
DRAGON_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Dragon')

DRAGON_FORMULA = FormulaRule(formula=[DRAGON_FORMULA_TEXT],
                              font=DRAGON_FONT,
                              border=DRAGON_BORDER,
                              fill=DRAGON_FILL)

DARKNESS_FILL = PatternFill(bgColor='3A3A3A', fill_type='solid')
DARKNESS_FONT = Font(color='FDFDFD')
DARKNESS_BORDER_COLOR = Color('BFBFBF', type='hex')
DARKNESS_BORDER_SIDE = Side(style='medium', color=DARKNESS_BORDER_COLOR, border_style='thin')
DARKNESS_BORDER = Border(left=DARKNESS_BORDER_SIDE,
                     right=DARKNESS_BORDER_SIDE,
                     top=DARKNESS_BORDER_SIDE,
                     bottom=DARKNESS_BORDER_SIDE)
DARKNESS_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Darkness')

DARKNESS_FORMULA = FormulaRule(formula=[DARKNESS_FORMULA_TEXT],
                              font=DARKNESS_FONT,
                              border=DARKNESS_BORDER,
                              fill=DARKNESS_FILL)

PSYCHIC_FILL = PatternFill(bgColor='DED1FF', fill_type='solid')
PSYCHIC_BORDER_COLOR = Color('A365D1', type='hex')
PSYCHIC_BORDER_SIDE = Side(style='medium', color=PSYCHIC_BORDER_COLOR, border_style='thin')
PSYCHIC_BORDER = Border(left=PSYCHIC_BORDER_SIDE,
                     right=PSYCHIC_BORDER_SIDE,
                     top=PSYCHIC_BORDER_SIDE,
                     bottom=PSYCHIC_BORDER_SIDE)
PSYCHIC_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Psychic')

PSYCHIC_FORMULA = FormulaRule(formula=[PSYCHIC_FORMULA_TEXT], border=PSYCHIC_BORDER, fill=PSYCHIC_FILL)

LIGHTNING_FILL = PatternFill(bgColor='FFFFCC', fill_type='solid')
LIGHTNING_BORDER_COLOR = Color('F6E400', type='hex')
LIGHTNING_BORDER_SIDE = Side(style='medium', color=LIGHTNING_BORDER_COLOR, border_style='thin')
LIGHTNING_BORDER = Border(left=LIGHTNING_BORDER_SIDE,
                     right=LIGHTNING_BORDER_SIDE,
                     top=LIGHTNING_BORDER_SIDE,
                     bottom=LIGHTNING_BORDER_SIDE)
LIGHTNING_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Lightning')

LIGHTNING_FORMULA = FormulaRule(formula=[LIGHTNING_FORMULA_TEXT], border=LIGHTNING_BORDER, fill=LIGHTNING_FILL)

WATER_FILL = PatternFill(bgColor='DDEBF7', fill_type='solid')
WATER_BORDER_COLOR = Color('8EA9DB', type='hex')
WATER_BORDER_SIDE = Side(style='medium', color=WATER_BORDER_COLOR, border_style='thin')
WATER_BORDER = Border(left=WATER_BORDER_SIDE,
                     right=WATER_BORDER_SIDE,
                     top=WATER_BORDER_SIDE,
                     bottom=WATER_BORDER_SIDE)
WATER_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Water')

WATER_FORMULA = FormulaRule(formula=[WATER_FORMULA_TEXT], border=WATER_BORDER, fill=WATER_FILL)

GRASS_FILL = PatternFill(bgColor='E2EFDA', fill_type='solid')
GRASS_BORDER_COLOR = Color('A9D08E', type='hex')
GRASS_BORDER_SIDE = Side(style='medium', color=GRASS_BORDER_COLOR, border_style='thin')
GRASS_BORDER = Border(left=GRASS_BORDER_SIDE,
                     right=GRASS_BORDER_SIDE,
                     top=GRASS_BORDER_SIDE,
                     bottom=GRASS_BORDER_SIDE)
GRASS_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Green')

GRASS_FORMULA = FormulaRule(formula=[GRASS_FORMULA_TEXT], border=GRASS_BORDER, fill=GRASS_FILL)

FIGHTING_FILL = PatternFill(bgColor='EBD5D1', fill_type='solid')
FIGHTING_BORDER_COLOR = Color('CD8A75', type='hex')
FIGHTING_BORDER_SIDE = Side(style='medium', color=FIGHTING_BORDER_COLOR, border_style='thin')
FIGHTING_BORDER = Border(left=FIGHTING_BORDER_SIDE,
                     right=FIGHTING_BORDER_SIDE,
                     top=FIGHTING_BORDER_SIDE,
                     bottom=FIGHTING_BORDER_SIDE)
FIGHTING_FORMULA_TEXT = make_formula_text(ENERGY_TYPE_COLUMN_LETTER, 'Fighting')

FIGHTING_FORMULA = FormulaRule(formula=[FIGHTING_FORMULA_TEXT], border=FIGHTING_BORDER, fill=FIGHTING_FILL)


# TODO FINISH THIS!!!
RED_FONT = Font(color='C00000')
NUMBER_OWNED_COLUMN_LETTER = get_number_owned_column_letter()
SHOW_LESS_THAN_FOUR_OWNED_FORMULA_TEXT = '=IF($H$2="Yes",4,0)'
SHOW_LESS_THAN_FOUR_OWNED_FORMULA = CellIsRule(operator='lessThan', stopIfTrue=False, font=RED_FONT)


def add_all_formatting_rules_to_sheet(ws: worksheet):
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, GRASS_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, FIRE_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, WATER_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, LIGHTNING_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, PSYCHIC_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, FIGHTING_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, DARKNESS_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, DRAGON_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, METAL_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, COLORLESS_FORMULA)
    ws.conditional_formatting.add(APPLIES_TO_ALL_ROWS, TRAINER_FORMULA)
